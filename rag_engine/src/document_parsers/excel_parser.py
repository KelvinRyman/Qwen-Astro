"""
Excel文档解析器

支持Microsoft Excel XLSX和XLS格式文档的文本提取、元数据提取和智能分块。
"""

import logging
from typing import List, Optional
from pathlib import Path
from .base_parser import DocumentParser, DocumentMetadata, DocumentChunk

try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


class ExcelParser(DocumentParser):
    """Excel文档解析器"""
    
    def __init__(self, config=None):
        super().__init__(config)
        self.max_rows_per_chunk = self.config.get('max_rows_per_chunk', 100)
        self.include_headers = self.config.get('include_headers', True)
        self.skip_empty_cells = self.config.get('skip_empty_cells', True)
        self.max_cell_length = self.config.get('max_cell_length', 1000)
    
    def can_parse(self, file_path: str) -> bool:
        """检查是否能解析Excel文件"""
        extension = Path(file_path).suffix.lower()
        
        if extension == '.xlsx' and OPENPYXL_AVAILABLE:
            return True
        elif extension == '.xls' and XLRD_AVAILABLE:
            return True
        
        return False
    
    def get_supported_extensions(self) -> List[str]:
        """获取支持的文件扩展名"""
        extensions = []
        if OPENPYXL_AVAILABLE:
            extensions.append('.xlsx')
        if XLRD_AVAILABLE:
            extensions.append('.xls')
        return extensions
    
    def extract_metadata(self, file_path: str) -> DocumentMetadata:
        """提取Excel元数据"""
        metadata = DocumentMetadata(format_type='excel')
        extension = Path(file_path).suffix.lower()
        
        try:
            if extension == '.xlsx' and OPENPYXL_AVAILABLE:
                metadata = self._extract_metadata_xlsx(file_path, metadata)
            elif extension == '.xls' and XLRD_AVAILABLE:
                metadata = self._extract_metadata_xls(file_path, metadata)
            
            # 文件大小
            metadata.file_size = Path(file_path).stat().st_size
            
        except Exception as e:
            self.logger.error(f"提取Excel元数据失败: {e}")
        
        return metadata
    
    def _extract_metadata_xlsx(self, file_path: str, metadata: DocumentMetadata) -> DocumentMetadata:
        """提取XLSX元数据"""
        workbook = load_workbook(file_path, read_only=True)
        
        # 基本信息
        sheet_names = workbook.sheetnames
        metadata.extra_metadata = {
            'sheet_count': len(sheet_names),
            'sheet_names': sheet_names
        }
        
        # 文档属性
        if hasattr(workbook, 'properties'):
            props = workbook.properties
            metadata.title = props.title
            metadata.author = props.creator
            metadata.subject = props.subject
            metadata.keywords = props.keywords
            
            if props.created:
                metadata.creation_date = props.created.strftime('%Y-%m-%d %H:%M:%S')
            if props.modified:
                metadata.modification_date = props.modified.strftime('%Y-%m-%d %H:%M:%S')
        
        # 统计信息
        total_rows = 0
        total_cols = 0
        
        for sheet_name in sheet_names:
            try:
                sheet = workbook[sheet_name]
                if sheet.max_row:
                    total_rows += sheet.max_row
                if sheet.max_column:
                    total_cols = max(total_cols, sheet.max_column)
            except Exception as e:
                self.logger.warning(f"读取工作表 {sheet_name} 失败: {e}")
        
        metadata.extra_metadata.update({
            'total_rows': total_rows,
            'max_columns': total_cols
        })
        
        workbook.close()
        return metadata
    
    def _extract_metadata_xls(self, file_path: str, metadata: DocumentMetadata) -> DocumentMetadata:
        """提取XLS元数据"""
        workbook = xlrd.open_workbook(file_path)
        
        # 基本信息
        sheet_names = workbook.sheet_names()
        metadata.extra_metadata = {
            'sheet_count': len(sheet_names),
            'sheet_names': sheet_names
        }
        
        # 统计信息
        total_rows = 0
        total_cols = 0
        
        for sheet_name in sheet_names:
            try:
                sheet = workbook.sheet_by_name(sheet_name)
                total_rows += sheet.nrows
                total_cols = max(total_cols, sheet.ncols)
            except Exception as e:
                self.logger.warning(f"读取工作表 {sheet_name} 失败: {e}")
        
        metadata.extra_metadata.update({
            'total_rows': total_rows,
            'max_columns': total_cols
        })
        
        return metadata
    
    def extract_text(self, file_path: str) -> str:
        """提取Excel全文"""
        extension = Path(file_path).suffix.lower()
        
        try:
            if extension == '.xlsx' and OPENPYXL_AVAILABLE:
                return self._extract_text_xlsx(file_path)
            elif extension == '.xls' and XLRD_AVAILABLE:
                return self._extract_text_xls(file_path)
            else:
                raise Exception("不支持的Excel格式或缺少相应的解析库")
        except Exception as e:
            self.logger.error(f"提取Excel文本失败: {e}")
            return ""
    
    def _extract_text_xlsx(self, file_path: str) -> str:
        """提取XLSX文本"""
        workbook = load_workbook(file_path, read_only=True)
        text_parts = []
        
        for sheet_name in workbook.sheetnames:
            try:
                sheet = workbook[sheet_name]
                sheet_text = self._extract_sheet_text_xlsx(sheet, sheet_name)
                if sheet_text:
                    text_parts.append(f"工作表: {sheet_name}\n{sheet_text}")
            except Exception as e:
                self.logger.warning(f"提取工作表 {sheet_name} 文本失败: {e}")
        
        workbook.close()
        return self._clean_text('\n\n'.join(text_parts))
    
    def _extract_text_xls(self, file_path: str) -> str:
        """提取XLS文本"""
        workbook = xlrd.open_workbook(file_path)
        text_parts = []
        
        for sheet_name in workbook.sheet_names():
            try:
                sheet = workbook.sheet_by_name(sheet_name)
                sheet_text = self._extract_sheet_text_xls(sheet, sheet_name)
                if sheet_text:
                    text_parts.append(f"工作表: {sheet_name}\n{sheet_text}")
            except Exception as e:
                self.logger.warning(f"提取工作表 {sheet_name} 文本失败: {e}")
        
        return self._clean_text('\n\n'.join(text_parts))
    
    def extract_chunks(self, file_path: str) -> List[DocumentChunk]:
        """提取Excel分块（按工作表分块）"""
        chunks = []
        base_metadata = {'file_name': Path(file_path).name, 'format_type': 'excel'}
        extension = Path(file_path).suffix.lower()
        
        try:
            if extension == '.xlsx' and OPENPYXL_AVAILABLE:
                chunks = self._extract_chunks_xlsx(file_path, base_metadata)
            elif extension == '.xls' and XLRD_AVAILABLE:
                chunks = self._extract_chunks_xls(file_path, base_metadata)
            
        except Exception as e:
            self.logger.error(f"提取Excel分块失败: {e}")
        
        return chunks
    
    def _extract_chunks_xlsx(self, file_path: str, base_metadata: dict) -> List[DocumentChunk]:
        """提取XLSX分块"""
        chunks = []
        workbook = load_workbook(file_path, read_only=True)
        
        for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
            try:
                sheet = workbook[sheet_name]
                sheet_text = self._extract_sheet_text_xlsx(sheet, sheet_name)
                
                if sheet_text:
                    chunk_metadata = self._create_chunk_metadata(
                        base_metadata,
                        {
                            'sheet_name': sheet_name,
                            'sheet_number': sheet_idx + 1,
                            'chunk_type': 'sheet',
                            'row_count': sheet.max_row or 0,
                            'column_count': sheet.max_column or 0
                        }
                    )
                    
                    chunk = DocumentChunk(
                        text=sheet_text,
                        metadata=chunk_metadata,
                        chunk_id=f"sheet_{sheet_idx + 1}",
                        section_title=sheet_name,
                        chunk_type='sheet'
                    )
                    chunks.append(chunk)
                    
            except Exception as e:
                self.logger.warning(f"处理工作表 {sheet_name} 失败: {e}")
        
        workbook.close()
        return chunks
    
    def _extract_chunks_xls(self, file_path: str, base_metadata: dict) -> List[DocumentChunk]:
        """提取XLS分块"""
        chunks = []
        workbook = xlrd.open_workbook(file_path)
        
        for sheet_idx, sheet_name in enumerate(workbook.sheet_names()):
            try:
                sheet = workbook.sheet_by_name(sheet_name)
                sheet_text = self._extract_sheet_text_xls(sheet, sheet_name)
                
                if sheet_text:
                    chunk_metadata = self._create_chunk_metadata(
                        base_metadata,
                        {
                            'sheet_name': sheet_name,
                            'sheet_number': sheet_idx + 1,
                            'chunk_type': 'sheet',
                            'row_count': sheet.nrows,
                            'column_count': sheet.ncols
                        }
                    )
                    
                    chunk = DocumentChunk(
                        text=sheet_text,
                        metadata=chunk_metadata,
                        chunk_id=f"sheet_{sheet_idx + 1}",
                        section_title=sheet_name,
                        chunk_type='sheet'
                    )
                    chunks.append(chunk)
                    
            except Exception as e:
                self.logger.warning(f"处理工作表 {sheet_name} 失败: {e}")
        
        return chunks
    
    def _extract_sheet_text_xlsx(self, sheet, sheet_name: str) -> str:
        """提取XLSX工作表文本"""
        rows_data = []
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx >= self.max_rows_per_chunk:
                break
            
            row_data = []
            for cell_value in row:
                if cell_value is not None:
                    cell_str = str(cell_value)
                    if len(cell_str) > self.max_cell_length:
                        cell_str = cell_str[:self.max_cell_length] + "..."
                    row_data.append(cell_str)
                elif not self.skip_empty_cells:
                    row_data.append("")
            
            if row_data:
                rows_data.append(" | ".join(row_data))
        
        return "\n".join(rows_data)
    
    def _extract_sheet_text_xls(self, sheet, sheet_name: str) -> str:
        """提取XLS工作表文本"""
        rows_data = []
        
        for row_idx in range(min(sheet.nrows, self.max_rows_per_chunk)):
            row_data = []
            for col_idx in range(sheet.ncols):
                try:
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    if cell_value:
                        cell_str = str(cell_value)
                        if len(cell_str) > self.max_cell_length:
                            cell_str = cell_str[:self.max_cell_length] + "..."
                        row_data.append(cell_str)
                    elif not self.skip_empty_cells:
                        row_data.append("")
                except Exception:
                    continue
            
            if row_data:
                rows_data.append(" | ".join(row_data))
        
        return "\n".join(rows_data)
